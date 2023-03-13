import socket
from kivy.uix.image import Image
from kivy.uix.stacklayout import StackLayout
from kivymd.app import MDApp
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.button import MDFlatButton, MDRaisedButton
from kivymd.uix.dialog import MDDialog
from kivymd.uix.screen import MDScreen
from kivymd.uix.screenmanager import MDScreenManager
from kivymd.uix.datatables import MDDataTable
from kivy.metrics import dp
from kivy.lang import Builder
from kivy.core.window import Window
from paypal import *
import random
import pandas as pd
from datetime import datetime
import time
import string
from barcode import EAN13
from barcode.writer import ImageWriter
import threading
import matplotlib.pyplot as plt
from kivy.garden.matplotlib.backend_kivyagg import FigureCanvasKivyAgg
import openpyxl
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore


SCAN_TIME = 5
Window.size = (300, 500)  # for mobile phones only

send_cond = False
recv_cond = False

pay_amount = 0.00  # amount generated in the phone
paid_amount = 0.00  # amount confirmed with payment gateway by the terminal
user_credentials = ''
pot_selected = 'None'
loyalty_name = 'None'
loyalty_code = 111111111111
unsorted = 0

#wb = openpyxl.load_workbook('data.xlsx')
cred = credentials.Certificate('firebase.json')
firebase_admin.initialize_app(cred)
db = firestore.client()

#trans_sheet = wb["transactions"]
#pots_sheet = wb["pots"]
#loyalty_sheet = wb["loyalty"]
#users_sheet = wb["users"]

#users_trans = pd.read_excel('data.xlsx', sheet_name='transactions')
#users_pots = pd.read_excel('data.xlsx', sheet_name='pots')
#users_loyalty = pd.read_excel('data.xlsx', sheet_name='loyalty')
#users_data = pd.read_excel('data.xlsx', sheet_name='users')

trans_sheet = db.collection(u'Transactions')
users_trans = pd.DataFrame(list(map(lambda x: x.to_dict(), list(trans_sheet.stream()))))

pots_sheet = db.collection(u'Pots')
users_pots = pd.DataFrame(list(map(lambda x: x.to_dict(), list(pots_sheet.stream()))))

loyalty_sheet = db.collection(u'Loyalty')
users_loyalty = pd.DataFrame(list(map(lambda x: x.to_dict(), list(loyalty_sheet.stream()))))

users_sheet = db.collection(u'Users')
users_data = pd.DataFrame(list(map(lambda x: x.to_dict(), list(users_sheet.stream()))))

users_names = users_data['name'].tolist()

pots = pd.DataFrame()
trans = pd.DataFrame()
loyalty = pd.DataFrame()

loyalty_names = []
pot_names = []
user_name = ''

TERMINAL_IP = socket.gethostbyname(socket.gethostname())
PORT_NUMBER = 5005
SIZE = 1024


def sender():
    global client_data, send_cond
    while True:
        try:
            while send_cond:
                send_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                send_socket.connect((TERMINAL_IP, PORT_NUMBER))
                print('connect')
                while True:
                    if client_data != '':
                        send_socket.send(bytes(client_data, 'utf-8'))
                        client_data = ''
        except:
            time.sleep(0.5)


thread_1 = threading.Thread(target=sender)


def receiver():
    global paid_amount, recv_cond

    while recv_cond:
        try:
            recv_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            recv_socket.bind((TERMINAL_IP, PORT_NUMBER))
            recv_socket.listen(10)
            print('binded')
            while True:
                listen_socket, address = recv_socket.accept()
                while True:
                    raw_msg = listen_socket.recv(1024)
                    paid_amount = raw_msg.decode('utf-8')
        except:
            time.sleep(0.5)
            print('binding...')


thread_2 = threading.Thread(target=receiver)


class WindowManager(MDScreenManager):
    pass


class LoadScreen(MDScreen):
    pass


class SignInScreen(MDScreen):
    def signin(self, username):
        global user_credentials, user_name, users_sheet, users_names, users_data
        #users_data = pd.read_excel('data.xlsx', sheet_name='users')
        users_data = pd.DataFrame(list(map(lambda x: x.to_dict(), list(users_sheet.stream()))))
        if username in users_names:
#            for row in users_sheet.iter_rows():
#                if username in row[0].value:
#                    user_row = row[0].row
#                    break
            user_credentials = users_sheet.document(f'{username}').get().to_dict()['client_id'] + \
                               users_sheet.document(f'{username}').get().to_dict()['client_secret']
            #user_credentials = users_sheet.cell(row=user_row, column=3).value + \
            #                   users_sheet.cell(row=user_row, column=4).value #3,4 are column numbers
            self.manager.current = 'home'
            user_name = username
            self.manager.ids.homescreen.load_pots()
            self.manager.ids.loyaltyscreen.load_loyalty()
        else:
            self.button = MDFlatButton(text='Try again', on_release=self.retry, pos_hint={'center_x': 0.5})
            self.dialog = MDDialog(title="SignIn status", text='Invalid username',
                                   size_hint=(0.7, 1), buttons=[self.button])
            self.dialog.open()

    def retry(self, obj):
        self.dialog.dismiss()
        self.ids.username.text = ''


class SignUpScreen(MDScreen):
    pass


class HomeScreen(MDScreen):

    pot_layout = StackLayout()

    def get_amount(self):
        global pay_amount, pot_selected
        pay_amount = "%.2f" % round(random.uniform(0.00, 9.99), 2)  # keeps 2 decimal points with trailing zeros
        self.ids.amount.text = "Amount: " + pay_amount + " GBP"
        self.manager.ids.justpayscreen.ids.paynow.disabled = False
        self.manager.ids.potpayscreen.ids.confirm.disabled = False
        self.manager.ids.potpayscreen.ids.potname.text = "None"
        self.manager.ids.potpaidscreen.ids.potname.text = "None"
        self.ids.justpay.disabled = False
        self.ids.loyalty.disabled = False
        self.clear_pots()
        self.manager.ids.loyaltyscreen.clear_loyalty()
        pot_selected = 'None'

    def load_data(self):
        global trans_sheet, pots_sheet, loyalty_sheet, users_sheet, trans, pots, loyalty, unsorted, wb, \
            users_pots, users_trans, users_loyalty, users_data, users_names, loyalty_names, pot_names, user_name

        #trans_sheet = wb["transactions"]
        #pots_sheet = wb["pots"]
        #loyalty_sheet = wb["loyalty"]
        #users_sheet = wb["users"]

        #users_trans = pd.read_excel('data.xlsx', sheet_name='transactions')
        #users_pots = pd.read_excel('data.xlsx', sheet_name='pots')
        #users_loyalty = pd.read_excel('data.xlsx', sheet_name='loyalty')
        #users_data = pd.read_excel('data.xlsx', sheet_name='users')

        trans_sheet = db.collection(u'Transactions')
        users_trans = pd.DataFrame(list(map(lambda x: x.to_dict(), list(trans_sheet.stream()))))

        pots_sheet = db.collection(u'Pots')
        users_pots = pd.DataFrame(list(map(lambda x: x.to_dict(), list(pots_sheet.stream()))))

        loyalty_sheet = db.collection(u'Loyalty')
        users_loyalty = pd.DataFrame(list(map(lambda x: x.to_dict(), list(loyalty_sheet.stream()))))

        users_sheet = db.collection(u'Users')
        users_data = pd.DataFrame(list(map(lambda x: x.to_dict(), list(users_sheet.stream()))))

        trans = users_trans[users_trans['username'] == user_name]
        pots = users_pots[users_pots['username'] == user_name]
        loyalty = users_loyalty[users_loyalty['username'] == user_name]

        pot_names = list(pots['name'])
        loyalty_names = list(loyalty['name'])
        users_names = users_data['name'].tolist()
        unsorted = trans[trans["pot"] == "None"].shape[0]

    def update_data(self):
        global pots_sheet, pot_names, wb
        for name in pot_names:
            pot_limit = pots[pots["name"] == name]["limit"].values[0]
            try:
                trans_spend = trans.groupby('pot').agg({'amount': ['sum', 'count']})
                pot_spend = trans_spend.loc[name][0]
                pot_weight = trans_spend.loc[name][1]
                pot_limit = pots[pots["name"] == name]["limit"].values[0]
                pot_usage = '{:.0%}'.format(float(pot_spend) / float(pot_limit))
            except:
                pot_spend = pot_limit - pots[pots["name"] == name]["balance"].values[0]
                pot_weight = pots[pots["name"] == name]["weight"].values[0]
                pot_usage = pots[pots["name"] == name]["usage"].values[0]
                pots_sheet.document(f'{user_name}-{name}').update({u'balance': pot_limit-pot_spend})
                pots_sheet.document(f'{user_name}-{name}').update({u'usage': pot_usage})
                pots_sheet.document(f'{user_name}-{name}').update({u'weight': int(pot_weight)})

            #for row in pots_sheet.iter_rows():
            #    if name in row[0].value and user_name in row[5].value:
            #        pots_sheet[f'C{row[0].row}'] = pot_limit - pot_spend
            #        pots_sheet[f'D{row[0].row}'] = pot_usage
            #        pots_sheet[f'E{row[0].row}'] = int(pot_weight)
            #        break

        #wb.save('data.xlsx')

    def load_pots(self):
        global unsorted, pot_names, user_name, users_pots, users_trans, trans, pots
        self.load_data()
        #pots.sort_values(by='Weight', ascending=False, inplace=True)
        self.update_sort_buttons()

        for name in pot_names:
            pot_limit = pots[pots["name"] == name]["limit"].values[0]
            try:
                trans_spend = trans.groupby('pot').agg({'amount': ['sum', 'count']})
                pot_spend = trans_spend.loc[name][0]
                pot_weight = trans_spend.loc[name][1]
                pot_usage = '{:.0%}'.format(float(pot_spend)/float(pot_limit))
            except: # in case there are no transactions for a certain pot
                pot_weight = pots[pots['name'] == name]['weight'].values[0]
                pot_usage = pots[pots['name'] == name]['usage'].values[0]

            button = MDRaisedButton(id=name, text=f'{name}-{pot_usage} ({int(pot_weight)})')
            button.bind(on_release=lambda btn=name: self.button_press(btn))
            button.disabled = True
            self.pot_layout.add_widget(button)
        self.add_widget(self.pot_layout)

    def unload_pots(self):
        try:
            while True: #as long as there are buttons in the layout
                self.pot_layout.remove_widget(self.pot_layout.children[0])
        except:
            pass
        self.remove_widget(self.pot_layout)

    def button_press(self, btn):
        global pay_amount, pot_selected, unsorted, pots
        self.load_data()
        pot_selected = btn.id
        pot_weight = pots[pots['name'] == pot_selected]['weight'].values[0]
        pot_usage = pots[pots['name'] == pot_selected]['usage'].values[0]
        self.manager.current = 'potpay'
        self.manager.ids.potpayscreen.ids.sort.text = f"Sort({unsorted})"
        self.manager.ids.potpayscreen.ids.amount.text = "Amount: " + str(pay_amount) + " GBP"
        self.manager.ids.justpayscreen.ids.amount.text = "Amount: " + str(pay_amount) + " GBP"
        self.manager.ids.potpayscreen.ids.potname.text = f'{pot_selected}-{pot_usage} ({int(pot_weight)})'
        self.manager.ids.potpaidscreen.ids.potname.text = f'{pot_selected}-{pot_usage} ({int(pot_weight)})'
        pot_limit = str(pots[pots["name"] == pot_selected]["limit"].values[0])
        self.manager.ids.potpayscreen.ids.potlimit.text = "limit\n" + str(round(float(pot_limit), 2))
        self.manager.ids.potpaidscreen.ids.potlimit.text = "limit\n" + str(round(float(pot_limit), 2))
        pot_balance = str(pots[pots["name"] == pot_selected]["balance"].values[0])
        self.manager.ids.potpayscreen.ids.potrem.text = "balance\n" + str(round(float(pot_balance), 2))
        self.manager.ids.potpaidscreen.ids.potrem.text = "balance\n" + str(round(float(pot_balance), 2))
        self.clear_pots()
        btn.disabled = True
        return btn.text

    def update_pot_buttons(self):
        global pots, pot_names, trans, user_name, wb
        self.load_data()
        #pots.sort_values(by='Weight', ascending=False, inplace=True)
        for i, name in enumerate(reversed(pot_names)):
            pot_usage = pots[pots["name"] == name]["usage"].values[0]
            pot_weight = int(pots[pots["name"] == name]["weight"].values[0])
            self.pot_layout.children[i].text = f'{name}-{pot_usage} ({int(pot_weight)})'

    def update_sort_buttons(self):
        global unsorted
        self.ids.sort.text = f'Sort({unsorted})'
        self.manager.ids.loyaltyscreen.ids.sort.text = f"Sort({unsorted})"
        self.manager.ids.justpayscreen.ids.sort.text = f"Sort({unsorted})"
        self.manager.ids.potpayscreen.ids.sort.text = f"Sort({unsorted})"
        self.manager.ids.potpaidscreen.ids.sort.text = f"Sort({unsorted})"

    def clear_pots(self):
        global pots, pot_names
        try:
            for i in range(len(pot_names)):
                self.pot_layout.children[i].disabled = False
        except:
            pass

    def disable_pots(self):
        global pots, pot_names
        try:
            for i in range(len(pot_names)):
                self.pot_layout.children[i].disabled = True
        except:
            pass


class JustPayScreen(MDScreen):

    def pay_now(self, btn):
        global pay_amount, paid_amount
        #self.send_pay()
        if self.pay():
            paid_amount = pay_amount
            self.record_pay()
            #btn.disabled = True
            self.manager.current = 'home'
            self.manager.ids.homescreen.update_pot_buttons()
            self.manager.ids.justpayscreen.ids.amount.text = "Amount: " + str(pay_amount) + " GBP"
            self.manager.ids.homescreen.ids.justpay.disabled = True
            self.manager.ids.homescreen.ids.loyalty.disabled = True
            self.manager.ids.potpayscreen.ids.confirm.disabled = True
            self.manager.ids.homescreen.disable_pots()
            pay_amount = 0.00
            self.manager.ids.homescreen.ids.amount.text = "Amount: " + str(pay_amount) + " GBP"
            self.manager.ids.homescreen.update_sort_buttons()

    def pay(self):
        global paid_amount, pay_amount, user_credentials
        payment_success = False
        client_id = user_credentials[:80]
        client_secret = user_credentials[80:]

        letters = string.ascii_lowercase
        batch_code = "".join(random.choice(letters) for i in range(200))
        #batch_code = '123' #inorder to make the payment fail
        response = make_payment(client_id, client_secret, pay_amount, batch_code)

        # Check the response to see if the payment was successful
        if response is None:
            response_text = "Payment failed without a response"
        elif response.status_code == 201:
            response_text = "Payment sent successfully!"
            payment_success = True
        else:
            response_text = f"Payment failed: {response.status_code}"

        close_button = MDFlatButton(text="Close", on_release=self.close_pay)

        self.dialog_pay = MDDialog(title="Payment status",
                               text=response_text,
                               size_hint=(0.7, 1), buttons=[close_button])
        self.dialog_pay.open()

        return payment_success

    def record_pay(self):
        global pay_amount, pot_selected, loyalty_code, unsorted, trans, pots, user_name, trans_sheet, users_pots, wb
        now = datetime.now()
        dt_string = now.strftime("%Y-%m-%d %H:%M:%S")
        lower, upper, numbers = string.ascii_lowercase, string.ascii_uppercase, '0123456789'
        chars = lower + upper + numbers
        trans_id = "".join(random.choice(chars) for i in range(10))
        try:
            #data_entry = [trans_id, pay_amount, pot_selected, loyalty_code, dt_string, user_name]
            #trans_sheet.append(data_entry)
            #wb.save('data.xlsx')

            record = { 'id': trans_id, 'amount': pay_amount, 'pot': pot_selected,
                          'loyalty': loyalty_code, 'timestamp': dt_string, 'username': user_name}
            doc_ref = db.collection(u'Transactions').document(record['id'])
            doc_ref.set(record)

            if pot_selected != "None":
                balance = pots_sheet.document(f'{user_name}-{pot_selected}').get().to_dict()['balance']
                weight = pots_sheet.document(f'{user_name}-{pot_selected}').get().to_dict()['weight']
                pots_sheet.document(f'{user_name}-{pot_selected}').update({u'balance': balance - float(pay_amount)})
                pots_sheet.document(f'{user_name}-{pot_selected}').update({u'weight': int(weight) + 1})

                #for row in pots_sheet.iter_rows():
                #    if pot_selected in row[0].value and user_name in row[5].value:
                #        user_pot_row = row[0].row
                #        pots_sheet[f'C{row[0].row}'] = pots_sheet[f'C{row[0].row}'].value - float(pay_amount)
                #        pots_sheet[f'E{row[0].row}'] = pots_sheet[f'E{row[0].row}'].value + 1
                #        break
                #wb.save('data.xlsx')
                self.manager.ids.homescreen.load_data()
                pot_limit = str(pots[pots["name"] == pot_selected]["limit"].values[0])
                self.manager.ids.potpaidscreen.ids.potlimit.text = "Limit\n" + str(round(float(pot_limit), 2))
                pot_balance = str(pots[pots["name"] == pot_selected]["balance"].values[0])
                self.manager.ids.potpaidscreen.ids.potrem.text = "Balance\n" + str(round(float(pot_balance), 2))
                pots_sheet.document(f'{user_name}-{pot_selected}').update({u'usage':
                                    '{:.0%}'.format(round((float(pot_limit)-float(pot_balance)) / float(pot_limit), 2))})
                #pots_sheet[f'D{user_pot_row}'] = '{:.0%}'.format(round((float(pot_limit)-float(pot_balance)) / float(pot_limit), 2))
                pot_usage = pots[pots["name"] == pot_selected]["usage"].values[0]
                pot_weight = int(pots[pots["name"] == pot_selected]["weight"].values[0])
                self.manager.ids.potpaidscreen.ids.potname.text = f'{pot_selected}-{pot_usage} ({pot_weight})'
                check_string = "Transaction added successfully.\n Spent " + pay_amount + \
                               " GBP\n from " + pot_selected + " Pot."
            else:
                self.manager.ids.homescreen.load_data()
                unsorted = trans[trans["pot"] == "None"].shape[0]
                check_string = "Transaction added successfully.\n Spent " + pay_amount + " GBP\n without choosing a pot"
        except:
            check_string = "Failed to update the transaction."
        close_button = MDFlatButton(text="Close", on_release=self.close_record,
                                    pos_hint={'center_x': 0.2})
        sort_button = MDRaisedButton(text=f"Sort ({unsorted}) ", on_release=self.goto_sort,
                                     pos_hint={'center_x': 0.8})

        self.dialog_record = MDDialog(title="Payment status", text = check_string,
                               size_hint=(0.7, 1), buttons=[close_button, sort_button])
        self.dialog_record.open()
        self.manager.ids.homescreen.update_data()
        #wb.save('data.xlsx')
        self.manager.ids.homescreen.update_pot_buttons()

    def send_pay(self):
        global user_credentials
        # Set up your PayPal API credentials
        client_id = "AfODQAuk-f-Ait4OKTHweC-hBhqIQXVkDaDDOgnrn8uNG9pS_tO9kfG71mevAyvIRKiLsD5d5KyMoZ09"
        client_secret = "EO0NHZ2yztKGLMcvEFLZw7dapQHHk6CKDfA2gYGw6aGK3rYNc7IbzgINrn8xbPEIcGJtl-ux7UXBvjLI"
        user_credentials = client_id+client_secret

    def close_record(self, obj):
        global pots, users_pots, pots_sheet, user_name
        self.dialog_record.dismiss()
        self.manager.ids.homescreen.update_data()

    def goto_sort(self, obj):
        self.manager.current = 'sort'
        self.manager.ids.sortscreen.show_table()
        self.dialog_record.dismiss()

    def close_pay(self, obj):
        global pots, users_pots, pots_sheet, user_name
        self.dialog_pay.dismiss()
        self.manager.ids.homescreen.update_data()


class PotPayScreen(MDScreen):

    def confirm_pay(self, btn):
        global pay_amount, paid_amount
        self.manager.current = 'potpaid'
        #self.manager.ids.justpayscreen.send_pay()
        if self.manager.ids.justpayscreen.pay():
            paid_amount = pay_amount
            self.manager.ids.justpayscreen.record_pay()
            #btn.disabled = True
            #self.manager.ids.potpaidscreen.amount.text = "Amount: " + str(pay_amount) + " GBP"
            self.manager.ids.homescreen.ids.justpay.disabled = True
            self.manager.ids.homescreen.ids.loyalty.disabled = True
            self.manager.ids.justpayscreen.ids.paynow.disabled = True
            self.manager.ids.homescreen.disable_pots()
            pay_amount = 0.00
            self.manager.ids.homescreen.ids.amount.text = "Amount: " + str(pay_amount) + " GBP"


class PotPaidScreen(MDScreen):
    pass


class LoyaltyScreen(MDScreen):

    loyalty_layout = StackLayout()

    def read_loyalty(self):
        self.manager.ids.homescreen.load_data()

    def clear_loyalty(self):
        global loyalty, loyalty_names, user_name, users_loyalty
        self.manager.ids.homescreen.load_data()
        try:
            for i in range(len(loyalty_names)):
                self.loyalty_layout.children[i].disabled = False
        except:
            pass

    def save_barcode(self, name, code):
        global user_name
        barcode = EAN13(code, writer=ImageWriter())
        barcode.save(f'barcodes/{user_name}/{name}')

    def load_loyalty(self):
        global loyalty, loyalty_names
        self.read_loyalty()
        for name in loyalty_names:
            button = MDRaisedButton(id=name, text=name)
            button.bind(on_release=lambda btn=name: self.button_press(btn))
            button.disabled = False
            self.loyalty_layout.add_widget(button)
        self.add_widget(self.loyalty_layout)

    def unload_loyalty(self):
        global loyalty_names

        try:
            for i in range(len(loyalty_names)):
                self.loyalty_layout.remove_widget(self.loyalty_layout.children[i])
        except:
            pass
        self.remove_widget(self.loyalty_layout)

    def button_press(self, btn):
        global loyalty_code, loyalty, users_loyalty, user_name
        self.manager.current = 'scan'
        self.manager.ids.scanscreen.ids.loyaltyname.text = btn.text + ' card'
        loyalty_code = str(int(loyalty[loyalty["name"] == btn.text]["code"].values[0]))
        self.save_barcode(btn.text, loyalty_code)
        self.manager.ids.scanscreen.ids.loyaltybarcode.source = f'barcodes/{user_name}/{btn.text}.png'
        self.clear_loyalty()
        btn.disabled = True
        return btn.text


class ScanScreen(MDScreen):
    def scan_loyalty(self):
        self.manager.ids.homescreen.ids.loyalty.disabled = True
        time.sleep(SCAN_TIME)
        self.manager.ids.homescreen.update_pot_buttons()
        self.manager.current = 'home'


class ShowScreen(MDScreen):
    def load_table(self):
        global trans, trans_sheet, users_trans, users_trans
        self.cols = trans.columns.values
        self.values = trans.values.tolist()
        layout = MDBoxLayout(orientation='vertical', spacing=0.01 * self.height)
        column_data = [(col, dp(10)) for col in self.cols]
        row_data = self.values
        self.table = MDDataTable(column_data=column_data, row_data=row_data, rows_num=100,
                                 size_hint=(0.9, 0.8), pos_hint={'center_x': 0.5, 'center_y': 0.5})
        self.button = MDFlatButton(text="Home", on_press=self.close_table)
        self.add_widget(self.table)
        self.add_widget(self.button)
        return layout

    def show_table(self):
        self.load_table()

    def close_table(self, obj):
        self.manager.current = 'home'
        self.manager.ids.homescreen.update_pot_buttons()


class ShowPotsScreen(MDScreen):
    def load_table(self):
        global pots, user_pots
        self.cols = pots.columns.values
        self.values = pots.values.tolist()

        layout = MDBoxLayout(orientation='vertical', spacing=0.01 * self.height)
        column_data = [(col, dp(15)) for col in self.cols]
        row_data = self.values
        self.table = MDDataTable(column_data=column_data, row_data=row_data,
                                 rows_num=100, size_hint=(0.9, 0.5), pos_hint={'center_x': 0.5, 'center_y': 0.7})
        self.button1 = MDFlatButton(text="Add a new pot", on_press=self.create_pot,
                                    pos_hint={'center_x': 0.2})
        self.button2 = MDFlatButton(text="Home", on_press=self.close_table,
                                    pos_hint={'center_x': 0.8})
        self.add_widget(self.table)
        self.add_widget(self.button1)
        self.add_widget(self.button2)
        self.pie_chart = FigureCanvasKivyAgg(self.create_pie_chart())
        self.pie_chart.print_png('pie_chart.png')
        self.pie_image = Image(source='pie_chart.png', pos_hint={'center_y': 0.3}, size_hint=(0.9, 0.5))
        self.add_widget(self.pie_image)
        return layout

    def create_pie_chart(self):
        global pots, pot_names, users_pots, pots_sheet, pots_sheet, user_name
        self.manager.ids.homescreen.load_data()
        values = pots['limit'].tolist()

        fig, ax = plt.subplots()
        ax.pie(values, labels=pot_names, autopct='%1.1f%%')
        return fig

    def show_table(self):
        self.load_table()

    def create_pot(self, obj):
        self.manager.current = 'newpot'

    def close_table(self, obj):
        self.manager.current = 'home'
        self.manager.ids.homescreen.update_pot_buttons()


class ShowNewScreen(MDScreen):
    pass


class SortScreen(MDScreen):
    selected_row_id = None
    num_cols = 5

    #sort_layout = MDBoxLayout(orientation='vertical')

    def get_unsorted(self):
        global trans
        self.manager.ids.homescreen.load_data()
        self.unsorted_trans = trans[trans["pot"] == "None"]
        self.unsorted_trans = self.unsorted_trans[['id', 'amount', 'timestamp']]
        self.cols = self.unsorted_trans.columns.values
        self.num_cols = len(self.cols)
        self.values = trans.values.tolist()
        self.unsorted_values = self.unsorted_trans.values.tolist()

    def load_table(self):
        self.get_unsorted()
        layout = MDBoxLayout(orientation='vertical', spacing=0.01*self.height)
        column_data = [(col, dp(10)) for col in self.cols]
        row_data = self.unsorted_values  # self.unsorted_values
        self.unsorted_table = MDDataTable(column_data=column_data, row_data=row_data, rows_num=100)
                                          #size_hint=(1.0, 0.9), pos_hint={'center_x': 0.5, 'center_y': 0.55})
        self.unsorted_table.bind(on_row_press=self.on_row_press)
        self.unsorted_table.bind(on_check_press=self.on_check_press)
        self.add_widget(self.unsorted_table)
        self.button = MDFlatButton(text="Home", on_press=self.close_table)
        #self.sort_layout.add_widget(self.button)
        self.add_widget(self.button)

    def show_table(self):
        self.load_table()

    def close_table(self, obj):
        self.manager.current = 'home'
        self.manager.ids.homescreen.update_pot_buttons()

    def on_row_press(self, instance_table, instance_row):
        global pots, pot_names
        self.selected_row_id = instance_row.table.row_data[instance_row.index//self.num_cols][0]
        close_button = MDFlatButton(text="Close", on_release=self.close_dialog)
        buttons = []
        for pot in pot_names:
            buttons.append(MDFlatButton(text=pot, on_release=self.assign_pot))
        buttons.append(close_button)
        i = 0
        buttons_list = []
        while i < len(buttons):
            buttons_list.append(buttons[i:i + 3]) # to create a nested list with 3 item lists
            i += 3

        self.dialog = MDDialog(title="Assign a pot",
                               text="Please choose a pot to assign to the selected transaction",
                               size_hint=(1, 0.5), buttons=buttons, type="simple")
        self.dialog.open()

    def assign_pot(self, btn):
        global unsorted, trans, pots, user_name, users_names, wb
        self.manager.ids.homescreen.load_data()
        sort_amount = trans.loc[trans['id'] == self.selected_row_id, 'amount'].values[0]

        trans_sheet.document(f'{self.selected_row_id}').update({u'pot': btn.text})

        #trans_sheet[f'C{assign_row}'] = btn.text
        #wb.save('data.xlsx')
        self.get_unsorted()
        pot_limit = pots[pots["name"] == btn.text]["limit"].values[0]
        balance = pots_sheet.document(f'{user_name}-{btn.text}').get().to_dict()['balance']
        weight = pots_sheet.document(f'{user_name}-{btn.text}').get().to_dict()['weight']
        pots_sheet.document(f'{user_name}-{btn.text}').update({u'balance': balance - float(sort_amount)})
        pots_sheet.document(f'{user_name}-{btn.text}').update({u'weight': int(weight) + 1})

        #for row in pots_sheet.iter_rows():
        #    if btn.text in row[0].value and user_name in row[5].value:
        #        user_pot_row = row[0].row
        #        pots_sheet[f'C{user_pot_row}'] = pots_sheet[f'C{user_pot_row}'].value - float(sort_amount)
        #        pots_sheet[f'E{user_pot_row}'] = pots_sheet[f'E{user_pot_row}'].value + 1
        #        break

        pot_balance = pots[pots["name"] == btn.text]["balance"].values[0]
        pots_sheet.document(f'{user_name}-{btn.text}').update({u'usage':
                            '{:.0%}'.format( round((float(pot_limit) - float(pot_balance)) / float(pot_limit), 2))})
        #pots_sheet[f'D{user_pot_row}'] = \
        #    '{:.0%}'.format( round((float(pot_limit) - float(pot_balance)) / float(pot_limit), 2))
        self.show_table()
        unsorted = trans[trans["pot"] == "None"].shape[0]
        self.manager.ids.homescreen.ids.sort.text = f"Sort({unsorted})"
        self.manager.ids.loyaltyscreen.ids.sort.text = f"Sort({unsorted})"
        self.manager.ids.justpayscreen.ids.sort.text = f"Sort({unsorted})"
        self.manager.ids.potpayscreen.ids.sort.text = f"Sort({unsorted})"
        self.manager.ids.potpaidscreen.ids.sort.text = f"Sort({unsorted})"
        self.dialog.dismiss()
        #wb.save('data.xlsx')
        self.manager.ids.homescreen.update_pot_buttons()

    def close_dialog(self, obj):
        self.dialog.dismiss()

    def on_check_press(self, instance_table, current_row):
        print(current_row)


class NewPotScreen(MDScreen):

    def load_new_pot(self, newpot):
        global pots, pot_names, user_name
        self.manager.ids.homescreen.load_data()
        if newpot in pot_names:
            button = MDRaisedButton(id=newpot, text=f'{newpot}-0% (0)')
            button.bind(on_release=lambda btn=newpot: self.manager.ids.homescreen.button_press(btn))
            button.disabled = False
            self.manager.ids.homescreen.pot_layout.add_widget(button)

    def add_pot(self, pot_name, pot_limit):
        global pots, pot_names, wb
        if pot_limit == '':
            check_string = f'Pot limit is empty!'
        else:
            try:
                #data_entry = [pot_name, pot_limit, pot_limit, "0%", 0, user_name]
                #pots_sheet.append(data_entry)
                check_string = f"{pot_name} pot added with a limit of {pot_limit} GBP successfully"
                record = {'name': pot_name, 'limit': pot_limit, 'balance': pot_limit,
                          'usage': "0%", 'weight': 0, 'username': user_name}
                doc_ref = pots_sheet.document(record['username']+'-'+record['name'])
                doc_ref.set(record)
            except:
                check_string = "Failed to add the pot"

        close_button = MDFlatButton(text="Close", on_release=self.close_dialog)

        self.dialog = MDDialog(title="Adding a new pot",
                               text= check_string,
                               size_hint=(0.7, 1), buttons=[close_button], disabled=True)
        self.dialog.open()
        #wb.save('data.xlsx')
        self.manager.ids.homescreen.load_data()

    def close_dialog(self, obj):
        self.manager.ids.homescreen.load_data()
        self.dialog.dismiss()
        self.manager.current = 'showpots'
        self.manager.ids.showpotsscreen.show_table()


class NewLoyaltyScreen(MDScreen):
    loyalty_layout = StackLayout()

    def load_new_loyalty(self, new):
        global loyalty, loyalty_names, user_name, loyalty_sheet, users_loyalty, users_loyalty
        self.manager.ids.homescreen.load_data()
        if new in loyalty_names:
            button = MDRaisedButton(id=new, text=new)
            button.bind(on_release=lambda btn=new: self.manager.ids.loyaltyscreen.button_press(btn))
            button.disabled = False
            self.manager.ids.loyaltyscreen.loyalty_layout.add_widget(button)

    def add_loyalty(self, loyalty_name, loyalty_code):
        global wb, loyalty_sheet, user_name
        if len(loyalty_code) != 12:
            check_string = "Invalid loyalty code!"
        else:
            try:
                #data_entry = [loyalty_name, loyalty_code, user_name]
                #loyalty_sheet.append(data_entry)
                check_string = f"{loyalty_name} loyalty card added successfully"
                record = {'name': loyalty_name, 'code': loyalty_code, 'username': user_name}
                doc_ref = loyalty_sheet.document(record['username'] + '-' + record['name'])
                doc_ref.set(record)
            except:
                check_string = "Failed to add the loyalty card"
                self.ids.confirm.disabled = True

        close_button = MDFlatButton(text="Close", on_release=self.close_dialog)

        self.dialog = MDDialog(title="Adding a new loyalty card",
                               text= check_string,
                               size_hint=(0.7, 1), buttons=[close_button])
        self.dialog.open()
        self.manager.ids.homescreen.load_data()
        #wb.save('data.xlsx')

    def close_dialog(self, obj):
        self.manager.ids.homescreen.load_data()
        self.dialog.dismiss()
        self.manager.current = 'loyalty'


class DemoApp(MDApp):
    def build(self):
        self.theme_cls.primary_palette = "Cyan"
        self.theme_cls.primary_hue = "900"
        self.theme_cls.theme_style = "Light"
        kv = Builder.load_file("helper.kv")
        return kv


if __name__ == "__main__":
    DemoApp().run()
